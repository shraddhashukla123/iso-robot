from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from iso_robot.helpers.slug import slugify

DEFAULT_POC_FILENAME = "ISO ROBOT RISK POC.xlsx"


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[4]


def default_poc_path() -> Path:
    return _repo_root() / DEFAULT_POC_FILENAME


def parse_risk_sources_sheet(path: Path) -> List[Dict[str, Any]]:
    """Parse **Risk Sources** rows into dicts (requires openpyxl)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Risk Sources"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    out: List[Dict[str, Any]] = []
    mode: Optional[str] = None  # 'numbered' | 'regional'

    for row in rows:
        if row is None:
            continue
        cells = list(row) + [None, None, None, None, None]
        c0, c1, c2, c3, c4 = cells[0], cells[1], cells[2], cells[3], cells[4]

        if c1 == "Source" and c2 == "Pull method":
            if c0 == "#":
                mode = "numbered"
            elif c0 == "Country / area":
                mode = "regional"
            continue

        if mode == "numbered":
            if c0 is None and c1 is None:
                continue
            if not c1:
                continue
            if isinstance(c0, str) and not str(c0).strip().isdigit() and c0 not in ("#",):
                continue
            name = str(c1).strip()
            if not name or name == "Source":
                continue
            pull = str(c2).strip() if c2 else ""
            monitor = str(c3).strip() if c3 else ""
            why = str(c4).strip() if c4 else ""
            src_id = slugify(name)
            out.append(
                {
                    "id": src_id,
                    "name": name,
                    "pull_method": pull,
                    "what_to_monitor": monitor,
                    "why_it_matters": why,
                    "region": None,
                }
            )

        elif mode == "regional":
            if not c1:
                continue
            country = str(c0).strip() if c0 else ""
            name = str(c1).strip()
            if not name or name == "Source":
                continue
            pull = str(c2).strip() if c2 else ""
            monitor = str(c3).strip() if c3 else ""
            why = str(c4).strip() if c4 else ""
            prefix = f"{country} " if country else ""
            label = f"{prefix}{name}".strip()
            src_id = slugify(label)[:96]
            out.append(
                {
                    "id": src_id,
                    "name": label,
                    "pull_method": pull,
                    "what_to_monitor": monitor,
                    "why_it_matters": why,
                    "region": country or None,
                }
            )

    return out


def _risk_domain_rows_from_sheet(path: Path) -> Tuple[List[str], List[Tuple[str, str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Risk Tables structure"]
        raw = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    header_idx = None
    for i, row in enumerate(raw):
        if row and str(row[0]).strip() == "Risk Domain":
            header_idx = i
            break
    if header_idx is None:
        return [], []

    data: List[Tuple[str, str, str]] = []
    for row in raw[header_idx + 1 :]:
        if not row or row[0] is None:
            continue
        dom = str(row[0]).strip()
        if not dom:
            continue
        why = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        examples = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if dom.lower().startswith("for middle east"):
            continue
        data.append((dom, why, examples))
        if len(data) >= 40:
            break
    return ["Risk Domain", "Why it matters", "Example sources"], data


def build_risk_library_seed_entries(path: Path) -> List[Dict[str, str]]:
    """Curated risk_library rows (~30–80) from **Risk Tables structure** + risk domain splits."""
    _, domain_rows = _risk_domain_rows_from_sheet(path)
    entries: List[Dict[str, str]] = []

    for dom, why, examples in domain_rows:
        tid = slugify(f"domain-{dom}")
        desc_parts = [why]
        if examples:
            desc_parts.append(f"Example sources: {examples}")
        entries.append(
            {
                "id": tid,
                "industry": "Maritime / Middle East shipping (POC)",
                "risk_domain": dom,
                "title": f"{dom} — core exposure",
                "description": " ".join(p for p in desc_parts if p).strip(),
                "tags": "poc-seed,tvra",
                "source_ref": "ISO ROBOT RISK POC.xlsx / Risk Tables structure",
                "notes": "",
            }
        )
        for part in [p.strip() for p in examples.replace(";", ",").split(",") if p.strip()]:
            if len(part) < 3:
                continue
            eid = slugify(f"{dom}-{part}")[:96]
            entries.append(
                {
                    "id": eid,
                    "industry": "Maritime / Middle East shipping (POC)",
                    "risk_domain": dom,
                    "title": f"{dom}: {part}",
                    "description": why,
                    "tags": "poc-seed,example-source",
                    "source_ref": "ISO ROBOT RISK POC.xlsx",
                    "notes": "",
                }
            )

    sources = parse_risk_sources_sheet(path)[:25]
    for s in sources:
        sid = slugify(f"source-{s['name']}")
        body = " ".join(
            x for x in (s.get("what_to_monitor"), s.get("why_it_matters")) if x
        ).strip()
        entries.append(
            {
                "id": sid,
                "industry": "External monitoring",
                "risk_domain": "Monitoring source",
                "title": s["name"],
                "description": body or (s.get("pull_method") or ""),
                "tags": "poc-seed,risk-source",
                "source_ref": "ISO ROBOT RISK POC.xlsx / Risk Sources",
                "notes": s.get("pull_method") or "",
            }
        )

    if len(entries) > 80:
        entries = entries[:80]
    return entries


def write_risk_library_csv(rows: List[Dict[str, str]], csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "id",
        "industry",
        "risk_domain",
        "title",
        "description",
        "tags",
        "source_ref",
        "notes",
    ]
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fieldnames})
